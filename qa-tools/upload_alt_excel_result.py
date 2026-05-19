from pathlib import Path
from datetime import datetime
import json
import re

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment


ALT_TEST_ROOT = Path(__file__).resolve().parent
ALLURE_RESULTS_DIR = ALT_TEST_ROOT / "alt-allure-results"
EXCEL_PATH = ALT_TEST_ROOT / "TC_Result.xlsx"


RESULT_COLUMNS = {
    "tc_id": "TC ID",
    "result": "실행결과",
    "status": "상태",
    "note": "비고",
    "executed_at": "실행일시",
}


def extract_tc_id(test_name: str):
    match = re.search(r"TC[_-]MJ[_-](\d{3})", test_name, re.IGNORECASE)

    if not match:
        return None

    return f"TC-MJ-{match.group(1)}"


def read_allure_results():
    if not ALLURE_RESULTS_DIR.exists():
        raise FileNotFoundError(f"Allure 결과 폴더를 찾을 수 없음: {ALLURE_RESULTS_DIR}")

    result_files = list(ALLURE_RESULTS_DIR.glob("*-result.json"))

    if not result_files:
        raise FileNotFoundError(f"Allure result json 파일이 없음: {ALLURE_RESULTS_DIR}")

    results = {}

    for file_path in result_files:
        data = json.loads(file_path.read_text(encoding="utf-8"))

        test_name = data.get("name", "")
        status = data.get("status", "unknown")
        message = data.get("statusDetails", {}).get("message", "")

        tc_id = extract_tc_id(test_name)

        if tc_id is None:
            continue

        if status == "passed":
            excel_result = "Pass"
            excel_status = "자동화 완료"
            note = "AltTester 자동화 테스트 통과"
        elif status == "failed":
            excel_result = "Fail"
            excel_status = "실패 확인 필요"
            note = message if message else "AltTester 자동화 테스트 실패"
        elif status == "broken":
            excel_result = "Broken"
            excel_status = "스크립트 확인 필요"
            note = message if message else "테스트 스크립트 또는 환경 오류"
        elif status == "skipped":
            excel_result = "Skipped"
            excel_status = "미실행"
            note = "테스트 미실행"
        else:
            excel_result = status
            excel_status = "확인 필요"
            note = "알 수 없는 테스트 상태"

        results[tc_id] = {
            "test_name": test_name,
            "result": excel_result,
            "status": excel_status,
            "note": note,
        }

    return results


def find_header_columns(sheet):
    headers = {}

    for cell in sheet[1]:
        if cell.value is None:
            continue

        header_name = str(cell.value).strip()

        for key, expected_name in RESULT_COLUMNS.items():
            if header_name == expected_name:
                headers[key] = cell.column

    missing = [
        name for key, name in RESULT_COLUMNS.items()
        if key not in headers
    ]

    if missing:
        raise RuntimeError(f"엑셀 컬럼 누락: {', '.join(missing)}")

    return headers


def apply_result_style(row_cells, result):
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    skip_fill = PatternFill("solid", fgColor="FFEB9C")
    default_fill = PatternFill("solid", fgColor="D9EAF7")

    if result == "Pass":
        fill = pass_fill
    elif result in ("Fail", "Broken"):
        fill = fail_fill
    elif result == "Skipped":
        fill = skip_fill
    else:
        fill = default_fill

    for cell in row_cells:
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def update_excel(results):
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"TC 엑셀 파일을 찾을 수 없음: {EXCEL_PATH}")

    workbook = load_workbook(EXCEL_PATH)
    sheet = workbook.active

    headers = find_header_columns(sheet)
    executed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    updated_count = 0
    not_found = []

    for tc_id, result_info in results.items():
        target_row = None

        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=headers["tc_id"]).value

            if cell_value is None:
                continue

            if str(cell_value).strip() == tc_id:
                target_row = row
                break

        if target_row is None:
            not_found.append(tc_id)
            continue

        sheet.cell(row=target_row, column=headers["result"]).value = result_info["result"]
        sheet.cell(row=target_row, column=headers["status"]).value = result_info["status"]
        sheet.cell(row=target_row, column=headers["note"]).value = result_info["note"]
        sheet.cell(row=target_row, column=headers["executed_at"]).value = executed_at

        row_cells = [
            sheet.cell(row=target_row, column=headers["result"]),
            sheet.cell(row=target_row, column=headers["status"]),
            sheet.cell(row=target_row, column=headers["note"]),
            sheet.cell(row=target_row, column=headers["executed_at"]),
        ]

        apply_result_style(row_cells, result_info["result"])
        updated_count += 1

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    workbook.save(EXCEL_PATH)

    return updated_count, not_found


def main():
    results = read_allure_results()
    updated_count, not_found = update_excel(results)

    print("Excel 업데이트 완료")
    print(f"업데이트 수: {updated_count}")
    print(f"엑셀 파일: {EXCEL_PATH}")

    if not_found:
        print(f"엑셀에서 찾지 못한 TC ID: {', '.join(not_found)}")


if __name__ == "__main__":
    main()