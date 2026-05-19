from pathlib import Path
from datetime import datetime
import json
import re

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


ALT_TEST_ROOT = Path(__file__).resolve().parent
ALLURE_RESULTS_DIR = ALT_TEST_ROOT / "alt-allure-results"
EXCEL_PATH = ALT_TEST_ROOT / "TC_Result.xlsx"

RESULT_COLUMNS = {
    "tc_id": "ID",
    "result": "실행결과",
    "status": "상태",
    "note": "비고",
}

HISTORY_SHEET_NAME = "History"


def get_run_info():
    now = datetime.now()
    return {
        "run_id": now.strftime("%Y%m%d_%H%M%S"),
        "executed_at": now.strftime("%Y-%m-%d %H:%M:%S"),
    }


def extract_tc_id(test_name: str):
    match = re.search(r"TC[_-]MJ[_-](\d{3})", test_name, re.IGNORECASE)

    if not match:
        return None

    return f"TC-MJ-{match.group(1)}"


def read_allure_results(run_info):
    if not ALLURE_RESULTS_DIR.exists():
        raise FileNotFoundError(f"Allure 결과 폴더를 찾을 수 없음: {ALLURE_RESULTS_DIR}")

    result_files = list(ALLURE_RESULTS_DIR.glob("*-result.json"))

    if not result_files:
        raise FileNotFoundError(f"Allure result json 파일이 없음: {ALLURE_RESULTS_DIR}")

    results = {}

    for file_path in result_files:
        data = json.loads(file_path.read_text(encoding="utf-8"))

        test_name = data.get("name", "")
        status = data.get("status", "unknown")
        message = data.get("statusDetails", {}).get("message", "")
        trace = data.get("statusDetails", {}).get("trace", "")

        tc_id = extract_tc_id(test_name)

        if tc_id is None:
            continue

        if status == "passed":
            excel_result = "Pass"
            excel_status = "자동화 완료"
            note = f"AltTester 자동화 테스트 통과 / 실행일시: {run_info['executed_at']}"
        elif status == "failed":
            excel_result = "Fail"
            excel_status = "실패 확인 필요"
            note = message if message else "AltTester 자동화 테스트 실패"
        elif status == "broken":
            excel_result = "Broken"
            excel_status = "스크립트 확인 필요"
            note = message if message else "테스트 스크립트 또는 환경 오류"
        elif status == "skipped":
            excel_result = "Skipped"
            excel_status = "미실행"
            note = "테스트 미실행"
        else:
            excel_result = status
            excel_status = "확인 필요"
            note = "알 수 없는 테스트 상태"

        results[tc_id] = {
            "tc_id": tc_id,
            "test_name": test_name,
            "allure_status": status,
            "result": excel_result,
            "status": excel_status,
            "note": note,
            "message": message,
            "trace": trace,
        }

    return results


def find_header_columns(sheet):
    headers = {}

    for cell in sheet[1]:
        if cell.value is None:
            continue

        header_name = str(cell.value).strip()

        for key, expected_name in RESULT_COLUMNS.items():
            if header_name == expected_name:
                headers[key] = cell.column

    missing = [
        name for key, name in RESULT_COLUMNS.items()
        if key not in headers
    ]

    if missing:
        raise RuntimeError(f"엑셀 컬럼 누락: {', '.join(missing)}")

    return headers


def apply_result_style(row_cells, result):
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    skip_fill = PatternFill("solid", fgColor="FFEB9C")
    default_fill = PatternFill("solid", fgColor="D9EAF7")

    if result == "Pass":
        fill = pass_fill
    elif result in ("Fail", "Broken"):
        fill = fail_fill
    elif result == "Skipped":
        fill = skip_fill
    else:
        fill = default_fill

    for cell in row_cells:
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def update_latest_result_sheet(sheet, results):
    headers = find_header_columns(sheet)

    updated_count = 0
    not_found = []

    for tc_id, result_info in results.items():
        target_row = None

        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=headers["tc_id"]).value

            if cell_value is None:
                continue

            if str(cell_value).strip() == tc_id:
                target_row = row
                break

        if target_row is None:
            not_found.append(tc_id)
            continue

        sheet.cell(row=target_row, column=headers["result"]).value = result_info["result"]
        sheet.cell(row=target_row, column=headers["status"]).value = result_info["status"]
        sheet.cell(row=target_row, column=headers["note"]).value = result_info["note"]

        row_cells = [
            sheet.cell(row=target_row, column=headers["result"]),
            sheet.cell(row=target_row, column=headers["status"]),
            sheet.cell(row=target_row, column=headers["note"]),
        ]

        apply_result_style(row_cells, result_info["result"])
        updated_count += 1

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    return updated_count, not_found


def create_history_sheet_if_needed(workbook):
    if HISTORY_SHEET_NAME in workbook.sheetnames:
        return workbook[HISTORY_SHEET_NAME]

    sheet = workbook.create_sheet(HISTORY_SHEET_NAME)

    headers = [
        "Run ID",
        "실행일시",
        "TC ID",
        "테스트명",
        "실행결과",
        "상태",
        "Allure 상태",
        "비고",
    ]

    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_index)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    return sheet


def append_history(sheet, results, run_info):
    start_row = sheet.max_row + 1

    for row_offset, result_info in enumerate(sorted(results.values(), key=lambda item: item["tc_id"])):
        row = start_row + row_offset

        sheet.cell(row=row, column=1).value = run_info["run_id"]
        sheet.cell(row=row, column=2).value = run_info["executed_at"]
        sheet.cell(row=row, column=3).value = result_info["tc_id"]
        sheet.cell(row=row, column=4).value = result_info["test_name"]
        sheet.cell(row=row, column=5).value = result_info["result"]
        sheet.cell(row=row, column=6).value = result_info["status"]
        sheet.cell(row=row, column=7).value = result_info["allure_status"]
        sheet.cell(row=row, column=8).value = result_info["note"]

        row_cells = [sheet.cell(row=row, column=col) for col in range(1, 9)]
        apply_result_style(row_cells, result_info["result"])

    return len(results)


def format_history_sheet(sheet):
    widths = {
        1: 18,
        2: 22,
        3: 14,
        4: 55,
        5: 12,
        6: 18,
        7: 14,
        8: 70,
    }

    thin = Side(style="thin", color="D9D9D9")

    for col_index, width in widths.items():
        sheet.column_dimensions[get_column_letter(col_index)].width = width

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    sheet.freeze_panes = "A2"


def update_excel(results, run_info):
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"TC 엑셀 파일을 찾을 수 없음: {EXCEL_PATH}")

    workbook = load_workbook(EXCEL_PATH)

    latest_sheet = workbook.active
    updated_count, not_found = update_latest_result_sheet(latest_sheet, results)

    history_sheet = create_history_sheet_if_needed(workbook)
    history_count = append_history(history_sheet, results, run_info)
    format_history_sheet(history_sheet)

    workbook.save(EXCEL_PATH)

    return updated_count, history_count, not_found


def main():
    run_info = get_run_info()
    results = read_allure_results(run_info)
    updated_count, history_count, not_found = update_excel(results, run_info)

    print("Excel 업데이트 완료")
    print(f"Run ID: {run_info['run_id']}")
    print(f"실행일시: {run_info['executed_at']}")
    print(f"최신 결과 업데이트 수: {updated_count}")
    print(f"History 추가 수: {history_count}")
    print(f"엑셀 파일: {EXCEL_PATH}")

    if not_found:
        print(f"엑셀에서 찾지 못한 TC ID: {', '.join(not_found)}")


if __name__ == "__main__":
    main()